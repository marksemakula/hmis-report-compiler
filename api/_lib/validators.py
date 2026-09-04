"""Ingestion: parsing and validation of OPD (105) and IPD (108) upload files."""
import csv
import io
import re
from datetime import datetime, date

from .metadata import mapping
from .diagnosis_map import map_diagnosis
from . import periods


_IPD_INDEX = None


def ipd_diagnosis_index():
    """Pairs HMIS 108 Section-6 'Cases'/'Deaths' data elements under user-friendly keys.

    Accepts e.g. 'CD01' (from CD01a/CD01b), 'CV01a' (from CV01a1/CV01a2),
    the exact DE code, and legacy '_2019' variants.
    """
    global _IPD_INDEX
    if _IPD_INDEX is not None:
        return _IPD_INDEX
    m = mapping()
    supported_ccs = {
        m["categoryCombos"]["IPD_AGE04_5P_SEX"]["id"],   # Age(0-4, 5+Yrs) & Sex
        m["categoryCombos"]["IPD_AGE_SEX"]["id"],        # maternal age bands
        m["categoryCombos"]["NEONATAL_AGE"]["id"],       # Age(0-7, 8-28 days)
    }
    index = {}

    def add(key, kind, deid, cc, legacy):
        if not key:
            return
        entry = index.setdefault(key, {})
        # prefer current codes over legacy _2019 variants
        if kind in entry and not entry.get(kind + "_legacy", True) and legacy:
            return
        entry[kind] = deid
        entry[kind + "_cc"] = cc
        entry[kind + "_legacy"] = legacy

    for deid, info in m["dataElements"]["HMIS108"].items():
        cc, code, name = info["categoryCombo"], info["code"], info["name"]
        if cc not in supported_ccs or not code:
            continue
        legacy = code.endswith("_2019")
        if re.search(r"-\s*Cases\s*$", name):
            add(code, "cases", deid, cc, legacy)
            add(re.sub(r"a(_2019)?$", "", code), "cases", deid, cc, legacy)
            add(re.sub(r"([a-d])1(_2019)?$", r"\1", code), "cases", deid, cc, legacy)
        elif re.search(r"-\s*Deaths\s*$", name):
            add(code, "deaths", deid, cc, legacy)
            add(re.sub(r"b(_2019)?$", "", code), "deaths", deid, cc, legacy)
            add(re.sub(r"([a-d])2(_2019)?$", r"\1", code), "deaths", deid, cc, legacy)
    _IPD_INDEX = index
    return index


OPD_COLUMNS = ["PatientNo", "VisitDate", "Age", "AgeUnit", "Sex", "DiagnosisCode", "VisitType"]
IPD_COLUMNS = ["PatientNo", "AdmissionDate", "DischargeDate", "Age", "AgeUnit", "Sex",
               "Ward", "DiagnosisCode", "Outcome"]

SEX_VALUES = {"M": "Male", "MALE": "Male", "F": "Female", "FEMALE": "Female"}
VISIT_TYPES = {"NEW": "New", "RE-ATTENDANCE": "Re", "REATTENDANCE": "Re", "RE": "Re", "RETURN": "Re"}
OUTCOMES = {"DISCHARGE", "DISCHARGED", "DEATH", "DIED", "REFERRED", "ABSCONDED", "TRANSFERRED", ""}
AGE_UNITS = {"YEARS", "YRS", "Y", "MONTHS", "M", "DAYS", "D", ""}

# Aliases from common ward names to the DHIS2 Ward Type category options
WARD_ALIASES = {
    "male medical": "MaleMedical", "malemedical": "MaleMedical",
    "female medical": "FemaleMedical", "femalemedical": "FemaleMedical",
    "male surgical": "MaleSurgical", "malesurgical": "MaleSurgical",
    "female surgical": "FemaleSurgical", "femalesurgical": "FemaleSurgical",
    "paediatric": "Paediatrics", "paediatrics": "Paediatrics", "pediatrics": "Paediatrics", "children": "Paediatrics",
    "maternity": "Maternity_Obstetric", "obstetric": "Maternity_Obstetric", "maternity_obstetric": "Maternity_Obstetric",
    "gynaecology": "Gynaecology", "gynecology": "Gynaecology", "gyn": "Gynaecology",
    "emergency": "Emergency Ward", "emergency ward": "Emergency Ward", "a&e": "Emergency Ward", "casualty": "Emergency Ward",
    "icu": "Intensive Care Unit (ICU)", "intensive care": "Intensive Care Unit (ICU)", "intensive care unit": "Intensive Care Unit (ICU)", "intensive care unit (icu)": "Intensive Care Unit (ICU)",
    "neonatal": "Neonatal Unit", "neonatal unit": "Neonatal Unit", "nicu": "Neonatal Unit",
    "tb": "TB", "tuberculosis": "TB",
    "psychiatric": "Psychiatric", "psychiatry": "Psychiatric", "mental health": "Psychiatric",
    "eye": "Eye", "ophthalmology": "Eye",
    "ent": "ENT",
    "orthopaedic": "Orthopaedic", "orthopedic": "Orthopaedic", "ortho": "Orthopaedic",
    "nutrition": "Nutrition",
    "palliative": "Palliative",
    "rehabilitation": "Rehabilitation Ward", "rehabilitation ward": "Rehabilitation Ward", "rehab": "Rehabilitation Ward",
    "acute care": "AcuteCareUnit", "acute care unit": "AcuteCareUnit", "acutecareunit": "AcuteCareUnit",
    "other": "Other wards", "other wards": "Other wards",
}


def normalise_ward(value: str):
    return WARD_ALIASES.get(str(value).strip().lower())


def _parse_date(value):
    value = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d",
                "%d/%m/%y", "%Y%m%d"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    # Excel serial number
    try:
        serial = float(value)
        if 20000 < serial < 60000:
            return date(1899, 12, 30) + __import__("datetime").timedelta(days=int(serial))
    except ValueError:
        pass
    return None


def age_in_years(age, unit):
    try:
        age = float(age)
    except (TypeError, ValueError):
        return None
    unit = str(unit or "years").strip().upper()
    if unit.startswith("D"):
        return age / 365.0
    if unit.startswith("M") and unit != "MALE":
        return age / 12.0
    return age


def _row_is_blank(values) -> bool:
    return all(v is None or str(v).strip() == "" for v in values)


def _pick_sheet(wb, expected_columns):
    """Choose the worksheet whose header row best matches the expected columns.

    Guards against workbooks where the active sheet is empty or unrelated
    (e.g. an export whose first sheet holds only formatting).
    """
    best, best_hits = None, 0
    for ws in wb.worksheets:
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first:
            continue
        header = {str(c).strip() for c in first if c is not None}
        hits = len(header & set(expected_columns))
        if hits > best_hits:
            best, best_hits = ws, hits
    return best if best is not None and best_hits >= 3 else None


# --------------------------------------------------- reading a column by name
# Every export spells the same column its own way - PatientNo, Patient No,
# Patient Number - and the difference is only ever whitespace and case. Matching
# on the literal string made a file that was right in substance fail on a space,
# so columns are compared with the spacing and case taken out.
def _key(name) -> str:
    return re.sub(r"\s+", "", str(name or "")).lower()


def _cells(row) -> dict:
    """One row keyed by normalised column name, so 'Patient No' and 'PatientNo'
    are the same column however the export spelled it."""
    return {_key(k): v for k, v in (row or {}).items() if k}


def _first(cells: dict, names):
    """The first of these column names the row actually carries, or ''."""
    for name in names:
        value = cells.get(_key(name))
        if value not in (None, ""):
            return value
    return ""


# --- Raw EMR export adapter -------------------------------------------------
# Jinja RRH's EMR exports one row per billed item (drug/test/service), so a
# single OPD visit spans many rows and uses different column names than the
# HMIS template. When those signature columns are present we collapse the
# export to one row per Visit No and rename columns to the template schema.
# The published clean template lacks these columns, so it passes through
# untouched.
#
# The EMR does not name those columns the same way twice: its visit report
# writes "All Diagnosis" and "Visit Category" where its diagnosis report writes
# "Diagnosis" and "Visit Type". Matching one spelling meant the second export
# fell straight through the adapter and then failed every required-field check
# on every one of its ten thousand rows - five identical complaints per line
# about columns the file had all along under other names. So the adapter looks
# for the ROLE a column plays and accepts any spelling the EMR gives it.
_EMR_COLUMNS = {
    "VisitNo":       ["Visit No", "Visit Number"],
    "PatientNo":     ["Patient No", "Patient Number", "Patient ID"],
    "Sex":           ["Gender", "Sex"],
    "Age":           ["Age"],
    "VisitDate":     ["Visit Date", "Date of Visit"],
    "VisitType":     ["Visit Category", "Visit Type", "Attendance Type"],
    # Free text, not the EMR's own disease code: the 105 mapping is built and
    # tested against the condition's name. "Disease Code" is deliberately absent.
    "DiagnosisCode": ["All Diagnosis", "Diagnosis", "Diagnosis Name", "Disease"],
}

# A visit number alongside a named condition is unmistakably a raw export: the
# clean template has no visit number, the 033B tally is two columns of code and
# value, strata carries neither, and 108 is keyed on admission. Visit type is
# not required to recognise one - where it is missing every visit counts as New,
# which is what the collapse already assumes.
_RAW_EMR_MARKERS = ("VisitNo", "DiagnosisCode")
_EMR_SEX = {"Female": "F", "Male": "M", "F": "F", "M": "M"}
_EMR_VISIT_TYPE = {
    "Consultation": "New", "Self Request": "New", "Refferal": "New", "Referral": "New",
    "Follow up": "Re-attendance", "RTT - Return To Treatment": "Re-attendance",
    "Represented": "Re-attendance", "CDDP": "Re-attendance",
}


def _emr_clean_patient_no(value):
    text = str(value or "").strip()
    return text[:-2] if text.endswith(".0") else text


def _emr_format_date(value):
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value or "")[:10]


def _is_raw_emr(fieldnames):
    got = {_key(f) for f in (fieldnames or []) if f}
    return all(any(_key(name) in got for name in _EMR_COLUMNS[role])
               for role in _RAW_EMR_MARKERS)


def _collapse_emr_visits(dict_rows):
    """Reshape raw EMR billing lines into one row per DIAGNOSIS.

    The facility records more than one diagnosis for a single visit and the
    Ministry counts every one of them, so collapsing to a single diagnosis per
    visit - as this did previously - silently dropped every co-diagnosis.

    Attendance is a different question: a visit is one attendance however many
    conditions were recorded at it. Only the first row for a visit therefore
    carries count_attendance, and the compiler honours that flag, so the
    attendance rows stay right while the condition rows are complete.
    """
    visits, order = {}, []
    for row in dict_rows:
        cells = _cells(row)

        def column(role):
            return _first(cells, _EMR_COLUMNS[role])

        vno = str(column("VisitNo") or "").strip()
        key = vno if vno else "__row_%d" % len(order)
        diagnosis = str(column("DiagnosisCode") or "").strip()
        if key not in visits:
            age = column("Age")
            try:
                age = str(int(float(age))) if str(age).strip() not in ("", "None") else ""
            except (TypeError, ValueError):
                age = str(age or "")
            sex = str(column("Sex") or "").strip()
            visits[key] = {
                "PatientNo": _emr_clean_patient_no(column("PatientNo")),
                "Age": age,
                "AgeUnit": "Years",
                "Sex": _EMR_SEX.get(sex, sex),
                "VisitDate": _emr_format_date(column("VisitDate")),
                "VisitType": _EMR_VISIT_TYPE.get(
                    str(column("VisitType") or "").strip(), "New"),
                "_diagnoses": [],
            }
            order.append(key)
        if diagnosis and diagnosis not in visits[key]["_diagnoses"]:
            visits[key]["_diagnoses"].append(diagnosis)

    out = []
    for key in order:
        v = visits.pop(key)
        diagnoses = v.pop("_diagnoses") or [""]
        for i, d in enumerate(diagnoses):
            out.append({**v, "DiagnosisCode": d,
                        # One attendance per visit, however many conditions.
                        "CountAttendance": "1" if i == 0 else "0"})
    return out


def parse_file(filename: str, content: bytes, expected_columns=None):
    """Return list of row dicts from CSV or Excel content.

    Blank rows (all cells empty) are skipped: Excel files often carry
    thousands of formatted-but-empty ghost rows after data is deleted,
    which would otherwise all fail validation.
    """
    if filename.lower().endswith((".xlsx", ".xls", ".xlsm")):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = (_pick_sheet(wb, expected_columns) if expected_columns else None) or wb.active
        rows = [r for r in ws.iter_rows(values_only=True) if not _row_is_blank(r)]
        if not rows:
            return []
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        dict_rows = [dict(zip(header, [("" if c is None else str(c)) for c in r])) for r in rows[1:]]
        if _is_raw_emr(header):
            return _collapse_emr_visits(dict_rows)
        return dict_rows
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    dict_rows = [dict(r) for r in reader if not _row_is_blank(r.values())]
    if reader.fieldnames and _is_raw_emr(reader.fieldnames):
        return _collapse_emr_visits(dict_rows)
    return dict_rows


# ------------------------------------------------------ recognising the file
# Each report takes a different shape of file, and the shapes are unmistakable:
# 033B is a two-column tally, 105:01 and 108 are patient-level line lists, and a
# generated extraction script writes strata.
#
# Uploading one against the wrong report used to fail every required-field check
# on every row, so a single mis-click on the report selector produced seventeen
# identical lines reading "PatientNo is required; Age is required; Sex is
# required..." about a file that was never going to have those columns. The
# reader is told what is missing and never what is actually wrong.
#
# Recognising the shape first lets the upload say, once, what the file is and
# which report it belongs to. Signatures are the smallest set of columns that
# cannot appear in any other report's file, so a template with extra columns
# still matches.
FILE_SHAPES = [
    ("SURV",   {"code", "value"}),
    ("STRATA", {"diagnosis", "band", "sex", "visit", "n"}),
    ("IPD",    {"patientno", "admissiondate"}),
    ("OPD",    {"patientno", "visitdate"}),
]

# What the file IS, named so a person recognises it at a glance...
SHAPE_DESCRIPTION = {
    "SURV":   "a weekly 033B tally (its columns are Code and Value)",
    "STRATA": "an extraction-script strata file (diagnosis, band, sex, visit, n)",
    "IPD":    "a patient-level 108 inpatient extract (PatientNo, AdmissionDate, Ward)",
    "OPD":    "a patient-level 105:01 outpatient extract (PatientNo, VisitDate, "
              "DiagnosisCode)",
}

# ...and what a report expects, worded to follow its own name without repeating
# it: "105:01 takes a patient-level extract".
SHAPE_EXPECTED = {
    "SURV":   "a two-column tally of Code and Value",
    "STRATA": "a strata file of diagnosis, band, sex, visit and n",
    "IPD":    "a patient-level extract with PatientNo, AdmissionDate and Ward",
    "OPD":    "a patient-level extract with PatientNo, VisitDate and DiagnosisCode",
}

# The report each shape belongs to, and the shapes each report accepts. The
# first entry is what that report is built around; the rest are alternatives it
# also understands, which today means 105:01 taking either a line list or the
# pre-aggregated strata a generated script writes.
SHAPE_REPORT = {"SURV": "SURV", "STRATA": "OPD", "IPD": "IPD", "OPD": "OPD"}
REPORT_SHAPES = {"OPD": ["OPD", "STRATA"], "IPD": ["IPD"], "SURV": ["SURV"]}


def identify_shape(fieldnames):
    """Which report's upload format these columns belong to, or None when the
    columns match nothing we know, in which case the ordinary per-row messages
    are the more useful answer."""
    got = {_key(f) for f in (fieldnames or []) if f}
    for shape, signature in FILE_SHAPES:
        if signature.issubset(got):
            return shape
    return None


def period_hint(shape, rows):
    """The period the file itself says it covers, as a DHIS2 identifier.

    Worth computing because the mis-selection is usually of period as well as
    report: a week-35 tally uploaded as July 2026 needs both corrected, and the
    file already knows which week it is."""
    if shape == "SURV":
        for row in rows or []:
            cells = _cells(row)
            if str(cells.get("code") or "").strip().lower() != "_start_yyyymmdd":
                continue
            d = _parse_date(cells.get("value"))
            if d:
                iso_year, iso_week, _ = d.isocalendar()
                return f"{iso_year}W{iso_week}"
        return None
    field = "admissiondate" if shape == "IPD" else "visitdate"
    counts = {}
    for row in rows or []:
        d = _parse_date(_cells(row).get(field, ""))
        if d:
            key = f"{d.year}{d.month:02d}"
            counts[key] = counts.get(key, 0) + 1
    # The modal month, not the first: an extract may carry a stray date either
    # side of the period it was run for.
    return max(counts, key=counts.get) if counts else None


def shape_mismatch(report_type: str, rows: list):
    """The message for a file that belongs to a different report, or None.

    Validation answers "is this row usable"; it cannot answer "is this the right
    file", and asked the wrong question it repeats itself once per row. Deciding
    it here, before validation, means the answer is given once, names the report
    the file does belong to, and carries the period the file says it covers."""
    shape = identify_shape(rows[0].keys() if rows else None)
    accepted = REPORT_SHAPES.get(report_type)
    if not shape or not accepted or shape in accepted:
        return None
    reports = mapping().get("reportTypes", {})
    got = reports.get(report_type, {"short": report_type})
    wants = reports.get(SHAPE_REPORT[shape], {"short": SHAPE_REPORT[shape],
                                              "periodType": "Monthly"})
    hint = period_hint(shape, rows)
    when = (f", period {periods.describe(wants['periodType'], hint)}"
            if hint else f" and pick a {str(wants['periodType']).lower()} period")
    return (f"This file is {SHAPE_DESCRIPTION[shape]}, but {got['short']} was "
            f"selected, and {got['short']} takes {SHAPE_EXPECTED[accepted[0]]}. "
            f"Select {wants['short']}{when}, then upload it again.")


# Columns the template names one thing and the registers another. Spacing and
# case are handled for every column already, so only genuine synonyms are listed
# here - the words a clerk would call the same column by.
_COLUMN_SYNONYMS = {
    "PatientNo":     ["Patient Number", "Patient ID"],
    "Sex":           ["Gender"],
    "VisitDate":     ["Date of Visit"],
    "VisitType":     ["Visit Category", "Attendance Type"],
    "DiagnosisCode": ["Diagnosis", "All Diagnosis", "Diagnosis Name"],
    "AdmissionDate": ["Date of Admission"],
    "DischargeDate": ["Date of Discharge"],
    "Outcome":       ["Discharge Outcome"],
}


def _canonical_row(row: dict) -> dict:
    """One row under the template's own column names, whatever the file called
    them.

    An export that names its columns Patient No, Gender and Visit Date holds
    every field 105:01 asks for, and used to be told all five were missing, on
    every row. The file's own columns are kept as well as renamed, because the
    compilers downstream read columns beyond the required set."""
    out = {str(k).strip(): ("" if v is None else str(v).strip())
           for k, v in (row or {}).items() if k}
    cells = {_key(k): v for k, v in out.items()}
    for canonical in OPD_COLUMNS + IPD_COLUMNS + ["CountAttendance"]:
        if out.get(canonical):
            continue
        value = _first(cells, [canonical] + _COLUMN_SYNONYMS.get(canonical, []))
        if value != "":
            out[canonical] = value
    return out


def validate_rows(report_type: str, rows: list, period: str):
    """Validate parsed rows. Returns (clean_rows, errors)."""
    m = mapping()
    code_index = m["HMIS105_01_codeIndex"] if report_type == "OPD" else m["HMIS108_codeIndex"]
    year, month = int(period[:4]), int(period[4:])
    errors, clean = [], []

    required = ["PatientNo", "Age", "Sex", "DiagnosisCode"] + (
        ["VisitDate", "VisitType"] if report_type == "OPD" else ["AdmissionDate", "Ward"]
    )

    for i, row in enumerate(rows, start=2):  # header is line 1
        row = _canonical_row(row)
        problems = []

        for field in required:
            if not row.get(field):
                problems.append(f"{field} is required")

        sex = SEX_VALUES.get(row.get("Sex", "").upper())
        if row.get("Sex") and not sex:
            problems.append(f"Sex '{row.get('Sex')}' is not recognised (use M or F)")

        years = age_in_years(row.get("Age"), row.get("AgeUnit"))
        if row.get("Age") and years is None:
            problems.append(f"Age '{row.get('Age')}' is not a valid number")
        if years is not None and not (0 <= years <= 130):
            problems.append(f"Age {years:.1f} years is outside the acceptable range")

        raw_code = row.get("DiagnosisCode", "").strip()
        # Translate EMR free-text diagnosis names into HMIS 105 codes.
        # Already-valid codes pass through unchanged.
        code = map_diagnosis(raw_code, code_index) if (report_type == "OPD" and raw_code) else raw_code
        code_norm = re.sub(r"\s+", "", code)
        if code and report_type == "OPD" and code_norm not in code_index:
            problems.append(f"Diagnosis code '{code}' does not match any HMIS 105 data element")
        if code and report_type == "IPD":
            if code_norm not in ipd_diagnosis_index():
                problems.append(f"Diagnosis code '{code}' does not match any HMIS 108 Cases/Deaths data element")

        if report_type == "OPD":
            d = _parse_date(row.get("VisitDate", ""))
            if row.get("VisitDate") and d is None:
                problems.append(f"VisitDate '{row.get('VisitDate')}' is not a valid date")
            in_period = d is not None and d.year == year and d.month == month
            vt = VISIT_TYPES.get(row.get("VisitType", "").upper())
            if row.get("VisitType") and not vt:
                problems.append(f"VisitType '{row.get('VisitType')}' must be New or Re-attendance")
            parsed = {"visit_date": d.isoformat() if d else None, "visit_type": vt}
        else:
            adm = _parse_date(row.get("AdmissionDate", ""))
            dis = _parse_date(row.get("DischargeDate", "")) if row.get("DischargeDate") else None
            if row.get("AdmissionDate") and adm is None:
                problems.append(f"AdmissionDate '{row.get('AdmissionDate')}' is not a valid date")
            if row.get("DischargeDate") and dis is None:
                problems.append(f"DischargeDate '{row.get('DischargeDate')}' is not a valid date")
            if adm and dis and dis < adm:
                problems.append("DischargeDate is earlier than AdmissionDate")
            ward = normalise_ward(row.get("Ward", "")) if row.get("Ward") else None
            if row.get("Ward") and ward is None:
                problems.append(f"Ward '{row.get('Ward')}' is not recognised")
            outcome = row.get("Outcome", "").upper()
            if outcome and outcome not in OUTCOMES:
                problems.append(f"Outcome '{row.get('Outcome')}' is not recognised")
            in_period = adm is not None and adm.year == year and adm.month == month
            parsed = {
                "admission_date": adm.isoformat() if adm else None,
                "discharge_date": dis.isoformat() if dis else None,
                "ward": ward,
                "outcome": "Death" if outcome in ("DEATH", "DIED") else (outcome.title() or None),
            }

        if problems:
            errors.append({"line": i, "patient": row.get("PatientNo", ""), "problems": problems})
        else:
            clean.append({
                **row, **parsed,
                "sex": sex, "age_years": years,
                "diagnosis_code": code_norm,
                "in_period": in_period,
                # Absent on a clean template upload, where every row is a visit.
                "count_attendance": str(row.get("CountAttendance", "1")).strip() != "0",
            })
    return clean, errors
